from multiprocessing.sharedctypes import Value
from dotenv import load_dotenv
from re import search
from telethon import TelegramClient
from telethon.tl.functions.users import GetFullUserRequest
from openpyxl import Workbook, load_workbook
import asyncio
import os

load_dotenv()
api_id = os.environ.get("api-id")
api_hash = os.environ.get("api-hash")
client = TelegramClient('anon', api_id, api_hash)
excelFile = "teleparse.xlsx"
async def main():

  keyword = input("Please type in the keyword to search for: ")

  # check if file exists in current directory, if not then create new
  # TODO: choose save directory to load excel sheet from, maybe keep in cache if developing GUI version
  latestId = None
  if (os.path.isfile(excelFile)):
    wb = load_workbook(excelFile)
    if wb[keyword]['A1']:
      latestId = wb[keyword]['A1']
  else:
    wb = Workbook()
  # for each of the chats, grab the messages
  sheets = wb.sheetnames

  # Use min id as well
  messages = await client.get_messages(entity=None, min_id=latestId, search=keyword)#min_id=latestId,
  if (len(messages) > 0):
    # Check if there are message
    #numRows = 0
    if (keyword not in sheets):
      wb.create_sheet(keyword)
    #else:
      #numRows = len(tuple(wb[keyword].rows))

    # Shift all the rows in the file, the length of messages down
    #if numRows > 0:
      #wb[keyword].move_range("A1:C"+str(numRows), rows=len(messages))
    
    # Hack for now, delete all rows and add them again
    # TODO: change implementation to adding rows to the end as opposed to replacing it all
    # But does not seem likely due to Telethon implementation
    wb[keyword].delete_cols(1, 3)
    rowNum = 1
    # add results
    for message in messages:
      name = await client.get_entity(message.peer_id.user_id)
      wb[keyword].cell(row=rowNum, column=1, value = message.id)
      wb[keyword].cell(row=rowNum, column=2, value = name.first_name)
      wb[keyword].cell(row=rowNum, column=3, value = message.message)
      rowNum = rowNum+1

  print("Operation Completed")
  wb.save(excelFile)
with client:
  client.loop.run_until_complete(main())